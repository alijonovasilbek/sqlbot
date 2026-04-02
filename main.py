import asyncio
import io
import logging
import os
import re
import unicodedata
from pathlib import Path

from aiogram import Bot, Dispatcher, F
from aiogram.filters import CommandStart
from aiogram.types import BufferedInputFile, Document, Message
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO)


def load_env_file(path: str = ".env") -> None:
    env_path = Path(path)
    if not env_path.exists():
        return
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


load_env_file()
BOT_TOKEN = os.getenv("BOT_TOKEN")
if not BOT_TOKEN:
    raise RuntimeError("BOT_TOKEN topilmadi. .env faylga BOT_TOKEN=... yozing.")

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()


# ─── Yordamchi funksiyalar ───────────────────────────────────────────────────

def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value)).strip().lower()
    text = re.sub(r"\s+", " ", text)
    return "" if text in {"", "nan", "none"} else text


def parse_int(value: object) -> str | None:
    """Har qanday sonni butun son stringga o'giradi: 1668713.0 -> '1668713'"""
    text = normalize_text(value)
    if not text:
        return None
    cleaned = text.replace(" ", "").replace("\xa0", "").replace(",", ".")
    try:
        number = float(cleaned)
    except ValueError:
        return None
    if number != number:  # NaN
        return None
    return str(int(number)) if number.is_integer() else None


def parse_str(value: object) -> str | None:
    """Hisobraqam kabi katta sonlarni string sifatida oladi"""
    text = normalize_text(value)
    if not text:
        return None
    cleaned = text.replace(" ", "").replace("\xa0", "")
    # float ko'rinishidagi butun sonni tozalash: '15001000900473459200.0' -> '15001000900473459200'
    if "." in cleaned:
        try:
            cleaned = str(int(float(cleaned)))
        except (ValueError, OverflowError):
            pass
    return cleaned if cleaned not in ("nan", "none") else None


# ─── Excel o'qish: СВОД SHEET ────────────────────────────────────────────────

TARGET_SHEET = "СВОД"  # Har doim shu sheetdan o'qiladi


def load_target_sheet_rows(file_bytes: bytes) -> tuple[str, list[list[object]]]:
    """'СВОД' sheetni o'qiydi. Topilmasa xato chiqaradi."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)

    # Sheet nomini katta-kichik harfga qaramay qidirish
    ws = None
    for worksheet in wb.worksheets:
        if worksheet.title.strip().upper() == TARGET_SHEET.upper():
            ws = worksheet
            break

    if ws is None:
        available = ", ".join(wb.sheetnames)
        raise ValueError(
            f"'{TARGET_SHEET}' sheet topilmadi!\n"
            f"Fayldagi sheetlar: {available}"
        )

    rows: list[list[object]] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if ws.row_dimensions[row_idx].hidden:
            continue
        rows.append(list(row))

    return ws.title, rows


# ─── Ustunlarni aniqlash ─────────────────────────────────────────────────────

def find_header_columns(rows: list[list[object]]) -> tuple[dict[str, int] | None, int]:
    """Header qatoridan ustun indekslarini topadi."""
    keywords = {
        "loan_id":    ("loan id",),
        "hisobraqam": ("asosiy hisobraqam", "hisobraqam"),
        "yangi_bal":  ("yangi balans",),
        "col_min":    ("min",),
        "col_max":    ("max",),
    }
    for row_idx, row in enumerate(rows[:20]):
        found: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            cell_text = normalize_text(cell)
            if not cell_text:
                continue
            for key, variants in keywords.items():
                if key not in found and any(v in cell_text for v in variants):
                    found[key] = col_idx
        if set(keywords) <= found.keys():
            return found, row_idx
    return None, -1


def find_marker_columns(rows: list[list[object]], from_row: int) -> tuple[dict[str, int] | None, int]:
    """'3','5','6','7','8' raqamli marker qatoridan ustun indekslarini topadi."""
    marker_map = {"3": "loan_id", "5": "hisobraqam", "6": "yangi_bal", "7": "col_min", "8": "col_max"}
    search_to = min(len(rows), from_row + 10)
    for row_idx in range(from_row, search_to):
        found: dict[str, int] = {}
        for col_idx, cell in enumerate(rows[row_idx]):
            m = parse_int(cell)
            if m in marker_map and marker_map[m] not in found:
                found[marker_map[m]] = col_idx
        if len(found) == 5:
            return found, row_idx
    return None, -1


# ─── Ma'lumot qatorlarini olish ──────────────────────────────────────────────

def extract_data_rows(rows: list[list[object]]) -> list[tuple[str, str, str, str, str]]:
    # 1. Header qatorini topamiz
    header_cols, header_row = find_header_columns(rows)

    # 2. Marker qatorini topamiz (А,1,2,3... qatori)
    search_from = header_row + 1 if header_row >= 0 else 0
    marker_cols, marker_row = find_marker_columns(rows, search_from)

    # Ustun indekslarini aniqlaymiz
    columns = marker_cols or header_cols
    if columns is None:
        raise ValueError(
            "Ustunlar aniqlanmadi! "
            "Fayl tuzilishi noto'g'ri yoki ustun nomlari o'zgargan."
        )

    # Ma'lumotlar qayerdan boshlanadi
    data_start = (marker_row + 1) if marker_row >= 0 else (header_row + 1)
    max_col = max(columns.values())

    result: list[tuple[str, str, str, str, str]] = []

    for row in rows[data_start:]:
        if len(row) <= max_col:
            continue

        loan_id    = parse_int(row[columns["loan_id"]])
        hisobraqam = parse_str(row[columns["hisobraqam"]])
        yangi_bal  = parse_int(row[columns["yangi_bal"]])
        col_min    = parse_int(row[columns["col_min"]])
        col_max    = parse_int(row[columns["col_max"]])

        # Biror qiymat yo'q bo'lsa o'tkazib yuboramiz
        if not all((loan_id, hisobraqam, yangi_bal, col_min, col_max)):
            continue

        # Loan ID mantiqiy tekshiruv
        if int(loan_id) <= 100000:
            continue

        result.append((loan_id, hisobraqam, yangi_bal, col_min, col_max))

    return result


# ─── Asosiy konvertatsiya ────────────────────────────────────────────────────

def excel_to_sql(file_bytes: bytes) -> tuple[str, int, str]:
    """
    Returns: (sql_text, row_count, sheet_name)
    """
    sheet_name, rows = load_target_sheet_rows(file_bytes)
    data_rows = extract_data_rows(rows)

    if not data_rows:
        raise ValueError(
            f"'{sheet_name}' sheetda ma'lumot topilmadi.\n"
            "Fayl tuzilishini tekshiring."
        )

    lines = ["begin", "delete from ln_test1;"]
    for loan_id, hisobraqam, yangi_bal, col_min, col_max in data_rows:
        lines.append(
            f"insert into ln_test1 values "
            f"({loan_id},'{hisobraqam}','{yangi_bal}',{col_min},{col_max},0);"
        )
    lines += ["end;", "/"]

    return "\n".join(lines), len(data_rows), sheet_name


# ─── Telegram handlers ───────────────────────────────────────────────────────

@dp.message(CommandStart())
async def cmd_start(message: Message):
    await message.answer(
        "👋 <b>Salom!</b>\n\n"
        "📊 Excel fayl yuboring — men uni SQL formatga o'giraman.\n\n"
        "<b>Qaysi ustunlar olinadi:</b>\n"
        "  • D → <code>Loan ID</code>\n"
        "  • F → <code>Asosiy hisobraqam</code>\n"
        "  • G → <code>Yangi balans</code>\n"
        "  • H → <code>Min</code>\n"
        "  • I → <code>Max</code>\n\n"
        "⚠️ <i>Faqat birinchi sheet o'qiladi!</i>\n"
        "📎 <i>.xlsx yoki .xls fayl yuboring</i>",
        parse_mode="HTML",
    )


@dp.message(F.document)
async def handle_document(message: Message):
    doc: Document = message.document

    allowed_mime = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    )
    is_excel = doc.mime_type in allowed_mime or (doc.file_name or "").lower().endswith((".xlsx", ".xls"))
    if not is_excel:
        await message.answer("⚠️ Faqat <b>.xlsx</b> yoki <b>.xls</b> fayl yuboring!", parse_mode="HTML")
        return

    wait_msg = await message.answer("⏳ Qayta ishlanmoqda...")

    try:
        file = await bot.get_file(doc.file_id)
        downloaded = await bot.download_file(file.file_path)
        content = downloaded.read()

        sql_text, count, sheet_name = excel_to_sql(content)

        base_name = (doc.file_name or "result").rsplit(".", 1)[0]
        sql_filename = base_name + ".sql"
        sql_file = BufferedInputFile(sql_text.encode("utf-8"), filename=sql_filename)

        await wait_msg.delete()
        await message.answer_document(
            sql_file,
            caption=(
                f"✅ <b>Tayyor!</b>\n"
                f"📋 Sheet: <code>{sheet_name}</code>\n"
                f"📝 INSERT qatorlar: <b>{count}</b> ta\n"
                f"📄 <code>{sql_filename}</code>"
            ),
            parse_mode="HTML",
        )

    except ValueError as e:
        await wait_msg.delete()
        await message.answer(f"❌ <b>Xatolik:</b> {e}", parse_mode="HTML")

    except Exception as e:
        await wait_msg.delete()
        logging.exception("Kutilmagan xatolik")
        await message.answer(
            f"❌ Kutilmagan xatolik:\n<code>{e}</code>",
            parse_mode="HTML",
        )


@dp.message()
async def handle_other(message: Message):
    await message.answer("📎 Iltimos, Excel fayl (.xlsx) yuboring yoki /start bosing.")


async def main():
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())